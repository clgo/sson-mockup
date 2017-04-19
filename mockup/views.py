from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.conf import settings

from openpyxl import load_workbook
from openpyxl.utils.exceptions import SheetTitleException

import sys
import json
from collections import OrderedDict


# Create your views here.
def home(request):
	context = {'sample_data': ['First Data', 'Second Data', 'Third Data']}
	return render(request, 'mockup/home.html', context)
	
def page1(request):
	context = {'sample_data': ['First Data', 'Second Data', 'Third Data']}
	return render(request, 'mockup/page1.html', context)

def page2(request):
	# read xls file
	# assume xlxs file name is known, thus using hardcoded
	# assume xlxs only 1 worksheet, and dataformat is known 2 column with data starting from row 2
	
	dict_item = {}

	context = {'sample_data': dict_item}
	return render(request, 'mockup/page2.html', context)

def getchartdata(request):

	ordered_dict = OrderedDict()
	datafile = 'sample_data.xlsx'

	if request.is_ajax():		

		dict_item = {}

		try:
			xlxsfile = settings.MEDIA_ROOT + datafile
			wb = load_workbook(filename = xlxsfile)
			sheetnames = wb.get_sheet_names()
			sheet = wb.get_sheet_by_name(sheetnames[0])
			for row in range(2, sheet.max_row + 1):
				name 	= sheet['A' + str(row)].value
				value 	= sheet['B' + str(row)].value
				dict_item[name] = value
				ordered_dict[name] = value
		
		except SheetTitleException as e:
			errorMsg = 'No worksheet found'
		except:
			errorMsg = "Unexpected error:", sys.exc_info()[0]
			print(errorMsg)

	return JsonResponse(ordered_dict)